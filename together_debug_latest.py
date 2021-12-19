# todo 导入模块区
# 内置库(exe能跑）
import json
import time
from time import sleep
from urllib import parse
import os
import copy
import requests
import configparser
import shutil
# 外部库(exe不能跑）
import pdfplumber
print(1)
import pandas as pd
print(2)
import openpyxl
print(3)
# print("卧槽")
# a=input("等一下")

# todo spider.py
def get_adress(bank_name):
    url = "http://www.cninfo.com.cn/new/information/topSearch/detailOfQuery"
    data = {
        'keyWord': bank_name,
        'maxSecNum': 10,
        'maxListNum': 5,
    }
    hd = {
        'Host': 'www.cninfo.com.cn',
        'Origin': 'http://www.cninfo.com.cn',
        'Pragma': 'no-cache',
        'Accept-Encoding': 'gzip,deflate',
        'Connection': 'keep-alive',
        'Content-Length': '70',
        'User-Agent': 'Mozilla/5.0(Windows NT 10.0;Win64;x64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 75.0.3770.100Safari / 537.36',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json,text/plain,*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    }
    r = requests.post(url, headers=hd, data=data)
    # print(r.text)
    r = r.content
    m = str(r, encoding="utf-8")
    pk = json.loads(m)
    orgId = pk["keyBoardList"][0]["orgId"]  # 获取参数
    plate = pk["keyBoardList"][0]["plate"]
    code = pk["keyBoardList"][0]["code"]
    # print(orgId, plate, code)
    return orgId, plate, code


def download_PDF(url, file_name, bank):  # 下载pdf
    url = url
    r = requests.get(url)
    f = open(bank + "/" + file_name + ".pdf", "wb")
    f.write(r.content)


def get_PDF(orgId, plate, code, bank):
    url = "http://www.cninfo.com.cn/new/hisAnnouncement/query"
    data = {
        'stock': '{},{}'.format(code, orgId),
        'tabName': 'fulltext',
        'pageSize': 30,
        'pageNum': 1,
        'column': plate,
        'category': 'category_ndbg_szsh;',
        'plate': '',
        'seDate': '',
        'searchkey': '',
        'secid': '',
        'sortName': '',
        'sortType': '',
        'isHLtitle': 'true',
    }
    hd = {
        'Host': 'www.cninfo.com.cn',
        'Origin': 'http://www.cninfo.com.cn',
        'Pragma': 'no-cache',
        'Accept-Encoding': 'gzip,deflate',
        'Connection': 'keep-alive',
        # 'Content-Length': '216',
        'User-Agent': 'User-Agent:Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/533.20.25 (KHTML, like Gecko) Version/5.0.4 Safari/533.20.27',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json,text/plain,*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'X-Requested-With': 'XMLHttpRequest',
        # 'Cookie': cookies
    }
    data = parse.urlencode(data)
    # print(data)
    r = requests.post(url, headers=hd, data=data)
    # print(r.text)
    r = str(r.content, encoding="utf-8")
    r = json.loads(r)
    reports_list = r['announcements']
    for report in reports_list:
        # if '摘要' in report['announcementTitle'] or ("201" not in report['announcementTitle'] and "202" not in report['announcementTitle']):

        if '摘要' in report['announcementTitle'] or ("202" not in report['announcementTitle']):
            continue
        if 'H' in report['announcementTitle']:
            continue
        else:  # http://static.cninfo.com.cn/finalpage/2019-03-29/1205958883.PDF
            pdf_url = "http://static.cninfo.com.cn/" + report['adjunctUrl']
            file_name = report['announcementTitle']
            print("正在下载：" + pdf_url, "存放在当前目录：/" + bank + "/" + file_name)
            download_PDF(pdf_url, file_name, bank)
            sleep(2)


def get_data(bank_list, filepath):
    print(filepath)
    rm(filepath)  #rm会删除dist/together_2下的所有文件
    input("暂时不用rm")
    for bank in bank_list:
        os.mkdir(bank)
        orgId, plate, code = get_adress(bank)  # 调用自写spider库中的get_adrss()方法获取参数
        input("get_adress没问题")
        get_PDF(orgId, plate, code, bank)  # 调用自写spider库中的get_PDF()方法下载pdf
        input("get_PDF没问题")
        print("-----------下一家!----------")


# todo rem.py
def rm(filepath):   #删掉中文文件夹
    wenjianjia_name_list=[]
    for root, dirs, files in os.walk(filepath):
        wenjianjia_name_list=dirs
        break
    print("exe所在目录有哪些文件夹：", wenjianjia_name_list)
    input("注意，要删东西了")
    for name in wenjianjia_name_list:
        if name>= u'u4e00' :  #假设文件夹的名字是中文
            shutil.rmtree(os.path.join(filepath, name))  # 递归删除文件夹
    for root, dirs, files in os.walk(filepath):
        wenjianjia_name_list=dirs
        break
    print("然后exe所在目录有哪些文件夹：", wenjianjia_name_list)

# todo inputco.py
def input_co(all_co, sheet):
    wb = openpyxl.load_workbook(all_co)
    ws = wb[sheet]
    co_num = ws.max_row  # 公司数目
    co_list = []
    for i in range(co_num):
        co_list.append(ws["A{}".format(i + 1)].value)
    return co_list


# todo cal.py(暂时没有解决变量来源的问题）
def getFileName(filepath):
    file_list = []
    root_list = []
    print(filepath)
    s = os.walk(filepath)
    # print(s)
    for root, dirs, files in s:
        # for filespath in files:
        #   file_list.append(os.path.join(root,filespath))
        # print("root:", root)         #together_2文件夹的目录
        # print("root:", root.replace(filepath+"\\",""))
        # print("dirs:", dirs)     #together_2文件夹下的所有文件夹组成的列表[公司1，公司2】
        # print("files:",files)     #together_2文件夹下的所有文件组成的列表
        file_list.append(files)
        a=root.replace(filepath + "\\", "")
        root_list.append(a)
    # 中文字符检查
    while root_list[0][0] <= u'u4e00':
        file_list.pop(0)  # 去除第一项
        root_list.pop(0)  # 去除第一项
    print("根目录（公司名）：", root_list)
    print("具体文件（年份财报）：", file_list)
    # namedict=dict(zip(root_list,file_list))
    return root_list, file_list


def count_pdf(name, allwords):  # 输入文件名和词库，输出该文件词库中词出现总数
    all_num = len(allwords)
    allwordsvalue = [0 for i in range(all_num)]
    print(name)
    with pdfplumber.open(name) as pp:
        pages_of_pdf = len(pp.pages)
        print(name + "的页数：", pages_of_pdf)
        for i in range(pages_of_pdf):
            page = pp.pages[i]
            textdata = page.extract_text()
            # print(textdata)
            if textdata is None:  # 如果该年有数据
                res = "该pdf无法提出文字，暂缺数据"
                all = -1  # -1表示该年pdf是图，识别不出
            else:
                for words_ind in range(all_num):
                    if allwords[words_ind] in textdata:
                        allwordsvalue[words_ind] += 1
                # data = open("text.txt", "a",encoding='utf-8', errors='ignore')    #新建一个txt
                # data.write(textdata)      #将pdf内容写入txt
                # encoding = 'utf-8'
                res = dict(zip(allwords, allwordsvalue))
                all = sum(res.values())
    print("统计结果：", res, " ---→  词库词汇出现次数：", all)
    return all


def toexcel(root_list, file_list, countsall):  # 导出数据到excel
    list = [[], []]
    name = []
    for i in range(len(file_list)):
        for j in range(len(file_list[i])):
            year = file_list[i][j]
            count = countsall[i][j]
            list[0].append(year)
            list[1].append(count)
            name.append(root_list[i])
    test = pd.DataFrame(columns=name, data=list)
    test = pd.DataFrame(test.values.T, index=test.columns, columns=test.index)
    print(test)
    test.to_csv('rq.csv', encoding='gbk', mode='a', header=None)


def run(filepath, all_words, sheet_):
    print("-----------------我要开始跑啦！------------------")
    allwords = input_co(all_words, sheet_)
    # filepath="D:\工作软件\爬虫学习\妞妞论文爬虫"      #文件根目录
    root_list, file_list = getFileName(filepath)  # 得到公司名字、对应公司财报
    list_len = len(file_list)  # 记录了的公司数
    countsall = copy.deepcopy(file_list)  # 深复制，构造一个数据容器装每个公司每年财报数据
    for i in range(list_len):  # 遍历所有公司
        num_pdf = len(file_list[i])
        for j in range(num_pdf):  # 遍历某公司的年份
            print("-----------------这是一条分割线------------------")
            name_co = root_list[i]
            name_year = file_list[i][j]
            name = name_co + "\\" + name_year
            print("现在统计的是    {}    之    {}".format(name_co, name_year))
            all = count_pdf(name, allwords)
            countsall[i][j] = all
            print(countsall)
    print("-----------------终于跑完啦！------------------")
    input("run跑起来没问题，但是写入得瞧瞧")
    toexcel(root_list, file_list, countsall)  # 写入excel
    print("--------------- 写入excel成功！------------------")
    # 到写入是好的
    rm(filepath)


# todo readini.py
def read_ini(db):
    root_dir = os.getcwd()  #编译后为D:\工作软件\爬虫学习\Niuniu_paper\dist\together_2
    src_ab_position=root_dir.replace("\dist\\together_2","")
    print(root_dir)
    print(src_ab_position)
    cf = configparser.ConfigParser()
    cf.read("config.ini", encoding='utf-8')  # 拼接得到config.ini文件的路径，直接使用
    # cf.read('D:\\工作软件\\爬虫学习\\Niuniu_paper\\config.ini', encoding='utf-8')  # 拼接得到config.ini文件的路径，直接使用
    # secs = cf.sections()  # 获取文件中所有的section(一个配置文件中可以有多个配置，如数据库相关的配置，邮箱相关的配置，                        每个section由[]包裹，即[section])，并以列表的形式返回
    options = cf.options(db)  # 获取某个section名为Mysql-Database所对应的键
    # items = cf.items("Mysql-Database")  # 获取section名为Mysql-Database所对应的全部键值对
    config_list = []
    for i in options:
        config_list.append(cf.get(db, i))
    # filepath = cf.get(db, "filepath")
    # all_co = cf.get(db, "all_co")
    # sheet = cf.get(db, "sheet")
    # batch_nums = int(cf.get(db, "batch_nums"))
    # print(filepath)
    # print(all_co)
    # print(sheet)
    # print(batch_nums)
    print("-----------成功读取配置文件---------")
    return cf


# todo operation.py
def main():
    all_list = input_co(all_co, sheet) #所有公司
    input("导入公司名称没问题")
    # ---------------------------------------------到这没问题-------------------------
    co_nums = len(all_list)  # 共1364个公司
    batch = int(co_nums / batch_nums) + 1  # 分1364/10批——137批
    part_list = [[] for i in range(batch)]
    for i in range(batch):
        for j in range(batch_nums):
            if len(all_list) != 0:
                out = all_list.pop(0)
                part_list[i].append(out)
            else:
                break
    for co_part in part_list:
        input("准备下数据")
        get_data(co_part, filepath)
        run(filepath, all_words, sheet_)


if __name__ == '__main__':
    # 读取配置参数
    db = input("请输入ini文件的section名")

    cf = read_ini(db)
    input("成功读取ini文件")
    filepath = cf.get(db, "filepath")
    all_co = cf.get(db, "all_co")
    sheet = cf.get(db, "sheet")
    batch_nums = int(cf.get(db, "batch_nums"))
    all_words = cf.get(db, "all_words")
    sheet_ = cf.get(db, "sheet_")
    # 执行主函数
    main()
